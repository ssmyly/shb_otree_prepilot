"""
make_spreadsheet.py — build shb_earnings_by_speed.xlsx

Simulates the SHB counting game for three speed archetypes (Fast / Average /
Slow) plus a blended population, recording how many problems each participant
gets right in each scored round (Round 1 = calibration, Rounds 2-3 = work; the
unpaid practice round is excluded) and their final take-home pay. Writes raw
per-participant rows to one sheet per group and a formula-driven Summary.

Economic logic is copied verbatim from shb_experiment/__init__.py (class C).
Only the human-performance parameters are modeled (see Assumptions sheet).
"""

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ───────── app constants (from class C in __init__.py) ─────────
SCORE_MULT = 5
NOISE_LO, NOISE_HI = -10, 10
MU, KAPPA1, BETA = 50.0, 0.40, 0.5
THRESH_BASE, THRESH_WORK = 45, 55
COST  = {0: 0, 1: 10, 2: 24, 3: 48}
CELLS = {0: 30, 1: 25, 2: 20, 3: 16}
BASELINE_SIZES = [12, 20, 30]
BASE_PAY, TOKEN_ENDOW, PPU = 0, 100, 20   # 2026-06-10: base pay removed, allowance $2->$5
DURATION, MAX_TOKENS = 90, 3

# ───────── performance model (assumptions) ─────────
OVERHEAD = 2.3
ACC_MEAN, ACC_SD = 0.90, 0.05
ACC_GRID_PENALTY = 0.0015
ROUND_JITTER_SD = 0.12

CATEGORIES = {            # base counting speed, seconds per cell
    'Fast':    0.20,
    'Average': 0.26,
    'Slow':    0.42,
}

def play_round(per_cell_person, acc, sizes):
    per_cell = per_cell_person * random.lognormvariate(0.0, ROUND_JITTER_SD)
    t, nc = 0.0, 0
    while True:
        cells = sizes if isinstance(sizes, int) else random.choice(sizes)
        ptime = per_cell * cells + OVERHEAD
        if t + ptime > DURATION:
            break
        t += ptime
        a = max(0.50, acc - ACC_GRID_PENALTY * (cells - 16))
        if random.random() < a:
            nc += 1
    return nc

def wage_given_signal(signal, premium):
    return max(0.0, round(MU + KAPPA1 * (signal - MU) + premium, 2))

def expected_net(score_hat, cost, premium):
    tot = sum(wage_given_signal(score_hat + n, premium)
              for n in range(NOISE_LO, NOISE_HI + 1) if score_hat + n >= THRESH_WORK)
    return tot / (NOISE_HI - NOISE_LO + 1) - cost

def choose_tokens(per_cell, acc, bank_pts, premium):
    best_t, best_ev = 0, -1e9
    for t in range(MAX_TOKENS + 1):
        if COST[t] > bank_pts:
            continue
        nc_hat = round(acc * DURATION / (per_cell * CELLS[t] + OVERHEAD))
        ev = expected_net(nc_hat * SCORE_MULT, COST[t], premium)
        if ev > best_ev:
            best_ev, best_t = ev, t
    return best_t

def simulate(per_cell, condition):
    acc = min(max(random.gauss(ACC_MEAN, ACC_SD), 0.55), 0.99)
    payoff = 0.0
    past = []
    rounds_correct, tokens, hired = [], [], []

    # Round 1: calibration (varying grid, no tokens)
    nc0 = play_round(per_cell, acc, BASELINE_SIZES)
    sig0 = nc0 * SCORE_MULT + random.randint(NOISE_LO, NOISE_HI)
    w0 = round(MU + KAPPA1 * (sig0 - MU), 2) if sig0 >= THRESH_BASE else 0.0
    payoff += w0
    past.append(w0)
    rounds_correct.append(nc0)
    hired.append(1 if sig0 >= THRESH_BASE else 0)

    # Rounds 2-3: work
    for _ in range(2):
        bank = TOKEN_ENDOW + payoff
        prem = BETA * sum(w - MU for w in past) if condition == 'no_shb' else 0.0
        t = choose_tokens(per_cell, acc, bank, prem)
        nc = play_round(per_cell, acc, CELLS[t])
        sig = nc * SCORE_MULT + random.randint(NOISE_LO, NOISE_HI)
        hit = sig >= THRESH_WORK
        wage = wage_given_signal(sig, prem) if hit else 0.0
        payoff += wage - COST[t]
        past.append(wage)
        rounds_correct.append(nc)
        tokens.append(t)
        hired.append(1 if hit else 0)

    payoff += BASE_PAY + TOKEN_ENDOW
    take_home = payoff / PPU
    return rounds_correct, tokens, hired, take_home

def draw_population_speed():
    pc = 0.26 * random.lognormvariate(0.0, 0.40)
    return min(max(pc, 0.12), 2.0)

# ───────── generate raw data ─────────
random.seed(2024)
N = 3000
groups = {}     # name -> list of row dicts
for name, speed in CATEGORIES.items():
    rows = []
    for i in range(N):
        cond = 'shb' if i % 2 == 0 else 'no_shb'
        rc, tok, hd, th = simulate(speed, cond)
        rows.append((i + 1, cond, rc[0], rc[1], rc[2], tok[0], tok[1],
                     hd[0], hd[1], hd[2], round(th, 2)))
    groups[name] = rows

# blended population (mostly fast, some slow)
pop_rows = []
for i in range(N):
    cond = 'shb' if i % 2 == 0 else 'no_shb'
    rc, tok, hd, th = simulate(draw_population_speed(), cond)
    pop_rows.append((i + 1, cond, rc[0], rc[1], rc[2], tok[0], tok[1],
                     hd[0], hd[1], hd[2], round(th, 2)))
groups['Population'] = pop_rows

# ───────── build workbook ─────────
FONT = 'Arial'
wb = openpyxl.Workbook()
wb.calculation.fullCalcOnLoad = True

hdr_fill = PatternFill('solid', fgColor='2C3E50')
hdr_font = Font(name=FONT, bold=True, color='FFFFFF', size=11)
title_font = Font(name=FONT, bold=True, size=14)
sub_font = Font(name=FONT, italic=True, size=10, color='555555')
bold = Font(name=FONT, bold=True)
base = Font(name=FONT)
blue = Font(name=FONT, color='0000FF')           # input assumption
center = Alignment(horizontal='center')
thin = Side(style='thin', color='BBBBBB')
box = Border(left=thin, right=thin, top=thin, bottom=thin)

RAW_HEADERS = ['Participant', 'Condition', 'R1 correct (Calibration)',
               'R2 correct (Work)', 'R3 correct (Work)', 'Tokens R2', 'Tokens R3',
               'Hired R1', 'Hired R2', 'Hired R3', 'Take-home ($)',
               'Never hired (any rd)']

def write_raw(ws, rows):
    ws.append(RAW_HEADERS)
    for c in range(1, len(RAW_HEADERS) + 1):
        cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for r in rows:
        ws.append(list(r))
    # helper flag: 1 if not hired in ANY of the three scored rounds (H,I,J all 0)
    for i in range(2, len(rows) + 2):
        ws.cell(i, 12, f"=IF(SUM(H{i}:J{i})=0,1,0)")
    for col, w in zip('ABCDEFGHIJKL',
                      [11, 10, 22, 18, 18, 10, 10, 9, 9, 9, 13, 13]):
        ws.column_dimensions[col].width = w
    ws.cell(1, 11).number_format = '$#,##0.00'
    ws.freeze_panes = 'A2'

raw_sheet_name = {'Fast': 'Fast (sim)', 'Average': 'Average (sim)',
                  'Slow': 'Slow (sim)', 'Population': 'Population (sim)'}

# Summary first
summ = wb.active
summ.title = 'Summary'

# raw sheets
for name in ['Fast', 'Average', 'Slow', 'Population']:
    ws = wb.create_sheet(raw_sheet_name[name])
    write_raw(ws, groups[name])

# ───────── Summary sheet ─────────
last = N + 1   # last data row in raw sheets
def rng(sheet, col):
    return f"'{sheet}'!{col}2:{col}{last}"

summ['A1'] = 'SHB Counting Game — Performance & Earnings by Speed'
summ['A1'].font = title_font
summ['A2'] = ('Problems correct per scored round and take-home pay, by participant '
              'speed. Practice round excluded. N = {:,} simulated participants per group.'
              .format(N))
summ['A2'].font = sub_font
summ['A3'] = ('Round 1 = calibration (varying grid). Rounds 2-3 = work rounds '
              '(grid shrinks with training tokens). Pay = $5 fungible allowance + net wages (no base pay).')
summ['A3'].font = sub_font

# main table
hdr_row = 5
headers = ['Category', 'Speed\n(sec/cell)',
           'R1 avg', 'R1 min', 'R1 max',
           'R2 avg', 'R2 min', 'R2 max',
           'R3 avg', 'R3 min', 'R3 max',
           'Avg pay', 'Min pay', 'Max pay',
           'Typical low\n(p10)', 'Typical high\n(p90)',
           'Avg tokens\n/work rd', '% hired\n(work rds)', '% never\nhired']
for c, h in enumerate(headers, start=1):
    cell = summ.cell(hdr_row, c, h)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = box

row_order = [('Fast', 'Fast (sim)'), ('Average', 'Average (sim)'),
             ('Slow', 'Slow (sim)'),
             ('Overall population\n(mostly fast, some slow)', 'Population (sim)')]
speeds = {'Fast (sim)': CATEGORIES['Fast'], 'Average (sim)': CATEGORIES['Average'],
          'Slow (sim)': CATEGORIES['Slow'], 'Population (sim)': None}

r = hdr_row + 1
for label, sh in row_order:
    summ.cell(r, 1, label).font = bold
    summ.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='center')
    sp = speeds[sh]
    spcell = summ.cell(r, 2, sp if sp is not None else 'mixed')
    spcell.font = blue if sp is not None else base
    spcell.alignment = center
    # rounds: C(3)=R1, D(4)=R2, E(5)=R3 correct
    for i, col in enumerate(['C', 'D', 'E']):
        summ.cell(r, 3 + i*3, f"=ROUND(AVERAGE({rng(sh, col)}),1)")
        summ.cell(r, 4 + i*3, f"=MIN({rng(sh, col)})")
        summ.cell(r, 5 + i*3, f"=MAX({rng(sh, col)})")
    # pay K(11)
    summ.cell(r, 12, f"=AVERAGE({rng(sh,'K')})")
    summ.cell(r, 13, f"=MIN({rng(sh,'K')})")
    summ.cell(r, 14, f"=MAX({rng(sh,'K')})")
    summ.cell(r, 15, f"=PERCENTILE({rng(sh,'K')},0.1)")
    summ.cell(r, 16, f"=PERCENTILE({rng(sh,'K')},0.9)")
    # avg tokens over F:G, % hired over I:J, % never hired over L
    summ.cell(r, 17, f"=ROUND(AVERAGE('{sh}'!F2:G{last}),2)")
    summ.cell(r, 18, f"=AVERAGE('{sh}'!I2:J{last})")
    summ.cell(r, 19, f"=AVERAGE('{sh}'!L2:L{last})")
    for c in range(1, 20):
        summ.cell(r, c).border = box
        if c >= 3 and c != 17:
            summ.cell(r, c).alignment = center
        if not summ.cell(r, c).font.bold and c != 2:
            if summ.cell(r, c).font.color is None or summ.cell(r, c).font.color.rgb != 'FF0000FF':
                summ.cell(r, c).font = base
    for c in (12, 13, 14, 15, 16):
        summ.cell(r, c).number_format = '$#,##0.00'
    summ.cell(r, 18).number_format = '0%'
    summ.cell(r, 19).number_format = '0%'
    r += 1

# widths
for col, w in zip('ABCDEFGHIJKLMNOPQRS',
                  [26, 9, 7, 7, 7, 7, 7, 7, 7, 7, 7, 9, 9, 9, 10, 10, 10, 9, 9]):
    summ.column_dimensions[col].width = w
summ.row_dimensions[hdr_row].height = 30
summ.freeze_panes = 'A6'

# legend / note under table
note_r = r + 1
notes = [
    'How to read this:',
    '• "Rx correct" = number of counting problems answered correctly in that 90-second round (avg / min / max across simulated participants).',
    '• Rounds 2-3 grids shrink when training tokens are bought, so a faster/decision affects how many problems are attempted.',
    '• "Avg pay" is mean take-home in USD; "Typical low/high" (p10-p90) is the middle-80% range; Min/Max is the full simulated range.',
    '• "% never hired" = share never hired in ANY of the 3 scored rounds; if they bought no tokens they keep the full $5 allowance, but if they spent it on tokens and still missed they can finish near $0 (no guaranteed minimum).',
    '• Base pay was REMOVED. Pay = $5 fungible allowance + net wages. There is no guaranteed floor; fast players still top out around $13.',
    '• Blue = input assumption you can change (speed). All other numbers are live formulas over the raw simulated rows (see the (sim) tabs).',
    '• SHB vs No-SHB pay is within a few cents on average, so conditions are pooled here.',
]
for i, t in enumerate(notes):
    cell = summ.cell(note_r + i, 1, t)
    cell.font = bold if i == 0 else base
summ.cell(note_r, 1).font = Font(name=FONT, bold=True, size=11)

# ───────── Assumptions sheet ─────────
aw = wb.create_sheet('Assumptions')
aw['A1'] = 'Assumptions & Model Notes'; aw['A1'].font = title_font
arows = [
    ('', ''),
    ('GAME ECONOMICS (copied exactly from shb_experiment/__init__.py, class C)', ''),
    ('Score multiplier (score = correct x 5)', SCORE_MULT),
    ('Signal noise (uniform integer)', f'{NOISE_LO} to {NOISE_HI}'),
    ('Prior mean (mu)', MU),
    ('Kappa_1 (weight on signal)', KAPPA1),
    ('Beta (No-SHB history weight)', BETA),
    ('Hire threshold — calibration (signal >=)', THRESH_BASE),
    ('Hire threshold — work rounds (signal >=)', THRESH_WORK),
    ('Token cost (points): 0/1/2/3', '0 / 10 / 24 / 48'),
    ('Grid cells by tokens: 0/1/2/3', '30 / 25 / 20 / 16'),
    ('Calibration grid sizes (cells)', '12, 20, 30 (random each problem)'),
    ('Base pay (REMOVED 2026-06-10)', f'{BASE_PAY} pts = $0.00'),
    ('Token allowance (fungible)', f'{TOKEN_ENDOW} pts = $5.00'),
    ('Points per USD', f'{PPU} (1 pt = $0.05)'),
    ('Round duration', f'{DURATION} s'),
    ('', ''),
    ('PERFORMANCE MODEL (assumptions — would be pinned down by a pilot)', ''),
    ('Time per problem', 'per_cell x cells + overhead'),
    ('Overhead per problem (type + 0.7s feedback gap)', f'{OVERHEAD} s'),
    ('Base accuracy ~ Normal(mean, sd)', f'{ACC_MEAN}, {ACC_SD}'),
    ('Accuracy penalty per extra cell above 16', ACC_GRID_PENALTY),
    ('Round-to-round speed wobble (lognormal sdlog)', ROUND_JITTER_SD),
    ('', ''),
    ('SPEED ARCHETYPES (seconds per cell)', ''),
    ('Fast  (~25th percentile of population)', CATEGORIES['Fast']),
    ('Average (~median)', CATEGORIES['Average']),
    ('Slow  (~85th-90th percentile)', CATEGORIES['Slow']),
    ('Population: lognormal, median 0.26, sdlog 0.40 (right-skewed: most fast, some slow)', ''),
    ('', ''),
    ('Token choice', 'payoff-maximizing (rational) given own speed'),
    ('Participants per group', N),
    ('Random seed', 2024),
]
for i, (k, v) in enumerate(arows, start=2):
    aw.cell(i, 1, k); aw.cell(i, 2, v)
    if k and v == '':
        aw.cell(i, 1).font = Font(name=FONT, bold=True)
    else:
        aw.cell(i, 1).font = base; aw.cell(i, 2).font = base
aw.column_dimensions['A'].width = 56
aw.column_dimensions['B'].width = 36

import os
OUTFILE = os.environ.get('OUTFILE', 'shb_earnings_by_speed.xlsx')
wb.save(OUTFILE)

# console sanity check
import statistics as stt
print("group        R1    R2    R3    avg$    min$   max$")
for name in ['Fast', 'Average', 'Slow', 'Population']:
    rows = groups[name]
    r1 = stt.mean(x[2] for x in rows); r2 = stt.mean(x[3] for x in rows); r3 = stt.mean(x[4] for x in rows)
    pay = [x[10] for x in rows]
    print(f"{name:11s} {r1:4.1f}  {r2:4.1f}  {r3:4.1f}  ${stt.mean(pay):5.2f}  "
          f"${min(pay):4.2f}  ${max(pay):5.2f}")
print("saved", OUTFILE)
